import openpyxl

from django.views import View
from django.shortcuts import render
from django.http import HttpResponse
from ..helpers import (report_folder, months_keys,
                       sql_get_vehicle, sql_get_shovels, sql_get_bulls, sql_get_allstoppagesdet, sql_get_wellmetervalue,
                       insert_vehicle, insert_shovels, insert_bulls, insert_allstoppagesdet)


class GetReport(View):

    @staticmethod
    def get(request):
        return render(request, 'get_report_page.html', context={'months': months_keys})

    @staticmethod
    def post(request):
        date, shift = request.POST.get('date').split('-'), int(request.POST.get('shift'))
        result_query_vehicles = []
        result_query_shovels = []
        result_query_bulls = []
        result_query_allstoppagesdet = []
        result_query_wellmetervalue = []

        if shift == 1 or shift == 2:
            result_query_vehicles.append([sql_get_vehicle(f'{date[2]}/{date[1]}/{date[0]}', shift)
                                         .getvalue().fetchall(), shift])
            result_query_shovels.append([sql_get_shovels(f'{date[2]}/{date[1]}/{date[0]}', shift)
                                        .getvalue().fetchall(), shift])
            result_query_bulls.append([sql_get_bulls(f'{date[2]}/{date[1]}/{date[0]}', shift)
                                      .getvalue().fetchall(), shift])
            result_query_wellmetervalue.append([sql_get_wellmetervalue(f'{date[2]}/{date[1]}/{date[0]}', shift)
                                               .getvalue().fetchall(), shift])
        if shift == 0:
            for sh in range(1, 3):
                result_query_vehicles.append([sql_get_vehicle(f'{date[2]}/{date[1]}/{date[0]}', sh)
                                             .getvalue().fetchall(), sh])
                result_query_shovels.append([sql_get_shovels(f'{date[2]}/{date[1]}/{date[0]}', sh)
                                            .getvalue().fetchall(), sh])
                result_query_bulls.append([sql_get_bulls(f'{date[2]}/{date[1]}/{date[0]}', sh)
                                          .getvalue().fetchall(), sh])
                result_query_wellmetervalue.append([sql_get_wellmetervalue(f'{date[2]}/{date[1]}/{date[0]}', sh)
                                                   .getvalue().fetchall(), sh])
            result_query_allstoppagesdet.append(sql_get_allstoppagesdet(f'{date[2]}/{date[1]}/{date[0]}')
                                                .getvalue().fetchall())

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Report_{date[2]}.{date[1]}.{date[0]}.xlsx"'
        wb = openpyxl.load_workbook(f'{report_folder}report.xlsx')

        ws = wb[str('БелАЗы ')]
        for rows in result_query_vehicles:
            if len(rows[0]) > 0:
                list_rows = rows[0]
                for row in list_rows:
                    if row[0] == '30' and rows[1] == 1:
                        insert_vehicle(ws, 7, row)
                    if row[0] == '32' and rows[1] == 1:
                        insert_vehicle(ws, 8, row)
                    if row[0] == '35' and rows[1] == 1:
                        insert_vehicle(ws, 9, row)
                    if row[0] == '30' and rows[1] == 2:
                        insert_vehicle(ws, 18, row)
                    if row[0] == '32' and rows[1] == 2:
                        insert_vehicle(ws, 19, row)
                    if row[0] == '35' and rows[1] == 2:
                        insert_vehicle(ws, 20, row)

        ws = wb[str('Экскаваторы')]
        for rows in result_query_shovels:
            if len(rows[0]) > 0:
                half_rows = rows[0]
                for row in half_rows:
                    if row[0] == '43' and rows[1] == 1:
                        insert_shovels(ws, 14, row, 'rd')
                    if row[0] == '43' and rows[1] == 2:
                        insert_shovels(ws, 29, row, 'rd')
                    if row[0] == '183' and rows[1] == 1:
                        insert_shovels(ws, 13, row, 'rd')
                    if row[0] == '183' and rows[1] == 2:
                        insert_shovels(ws, 28, row, 'rd')
                    if row[0] == '283' and rows[1] == 1:
                        insert_shovels(ws, 12, row, 'rd')
                    if row[0] == '283' and rows[1] == 2:
                        insert_shovels(ws, 27, row, 'rd')
                    # if row[0] == '284' and rows[1] == 1:
                    #     insert_shovels(ws, 15, row, 'rd')
                    # if row[0] == '284' and rows[1] == 2:
                    #     insert_shovels(ws, 30, row, 'rd')
                    if row[0] == '285' and rows[1] == 1:
                        insert_shovels(ws, 16, row, 'rd')
                    if row[0] == '285' and rows[1] == 2:
                        insert_shovels(ws, 31, row, 'rd')
                    if row[0] == '58' and rows[1] == 1:
                        insert_shovels(ws, 11, row, 'ir')
                    if row[0] == '58' and rows[1] == 2:
                        insert_shovels(ws, 26, row, 'ir')
                    if row[0] == '45' and rows[1] == 1:
                        insert_shovels(ws, 10, row, 'ir')
                    if row[0] == '45' and rows[1] == 2:
                        insert_shovels(ws, 25, row, 'ir')
                    if row[0] == '286' and rows[1] == 1:
                        insert_shovels(ws, 17, row, 'rd')
                    if row[0] == '286' and rows[1] == 2:
                        insert_shovels(ws, 32, row, 'rd')

        ws = wb[str('Бульдозеры')]
        for rows in result_query_bulls:
            if len(rows[0]) > 0:
                half_rows = rows[0]
                for row in half_rows:
                    if row[0] == 9 and rows[1] == 1:
                        insert_bulls(ws, 8, row)
                    if row[0] == 9 and rows[1] == 2:
                        insert_bulls(ws, 12, row)
                    if row[0] == 186 and rows[1] == 1:
                        insert_bulls(ws, 7, row)
                    if row[0] == 186 and rows[1] == 2:
                        insert_bulls(ws, 11, row)
                    if row[0] == 4 and rows[1] == 1:
                        insert_bulls(ws, 9, row)
                    if row[0] == 4 and rows[1] == 2:
                        insert_bulls(ws, 13, row)
                    if row[0] == 15 and rows[1] == 1:
                        insert_bulls(ws, 10, row)
                    if row[0] == 15 and rows[1] == 2:
                        insert_bulls(ws, 14, row)

        ws = wb[str('Техника new')]
        for rows in result_query_allstoppagesdet:
            if len(rows) > 0:
                row = rows[0]
                if row[1] == 45:
                    insert_allstoppagesdet(ws, 9, row)
                if row[1] == 28:
                    insert_allstoppagesdet(ws, 12, row)
                if row[1] == 31:
                    insert_allstoppagesdet(ws, 13, row)
                if row[1] == 186:
                    insert_allstoppagesdet(ws, 15, row)
                if row[1] == 9:
                    insert_allstoppagesdet(ws, 16, row)
                if row[1] == 4:
                    insert_allstoppagesdet(ws, 18, row)
                if row[1] == 15:
                    insert_allstoppagesdet(ws, 19, row)
                if row[1] == 58:
                    insert_allstoppagesdet(ws, 21, row)
                if row[1] == 283:
                    insert_allstoppagesdet(ws, 22, row)
                if row[1] == 183:
                    insert_allstoppagesdet(ws, 23, row)
                if row[1] == 43:
                    insert_allstoppagesdet(ws, 24, row)
                if row[1] == 284:
                    insert_allstoppagesdet(ws, 25, row)
                if row[1] == 285:
                    insert_allstoppagesdet(ws, 26, row)
                if row[1] == 286:
                    insert_allstoppagesdet(ws, 27, row)

        ws = wb[str('Суточный рапорт')]
        ws['C2'] = f'{date[2]}.{date[1]}.{date[0]}'
        for rows in result_query_wellmetervalue:
            if len(rows[0]) > 0:
                list_rows = rows[0]
                for row in list_rows:
                    if row[0] == 'Зумпф №1 СВК Ю' and rows[1] == 1:
                        ws[f'AX137'] = row[1]
                    if row[0] == 'Зумпф №1 СВК Ю' and rows[1] == 2:
                        ws[f'CS137'] = row[1]
                    if row[0] == 'Расходомер Промплощадка' and rows[1] == 1:
                        ws[f'AX149'] = row[1]
                    if row[0] == 'Расходомер Промплощадка' and rows[1] == 2:
                        ws[f'CS149'] = row[1]

        wb.save(response)

        return response
