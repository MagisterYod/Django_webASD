from typing import Callable

import openpyxl

from django.views import View
from django.shortcuts import render
from django.http import HttpResponse
from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from ..helpers import report_folder, months_keys, p_date, sql_get_walking_report_crew_new


class GetReport17(View):

    @staticmethod
    def get(request):
        return render(request, 'get_report_17.html', context={'months': months_keys})

    @staticmethod
    def post(request):
        month = request.POST.get('month')

        td = Side(border_style='dotted')
        cell_dotted = NamedStyle(name='cell')
        cell_dotted.alignment = Alignment(horizontal="center", vertical="center")
        cell_dotted.border = Border(top=td, left=td, bottom=td, right=td)
        cell_dotted.font = Font(name='Times New Roman', size=12)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Report_Walking_{month}.{p_date[2]}.xlsx"'
        wb = openpyxl.load_workbook(f'{report_folder}ReportWalkingCrew.xlsx')
        ws = wb[str('Детализация')]
        ws['F3'] = f'{[key for key, val in months_keys if val == month][0]} {p_date[2]} г.'

        td = Side(border_style='dotted')
        cell = NamedStyle(name='cell')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=td, left=td, bottom=td, right=td)
        cell.font = Font(name='Times New Roman', size=12)
        none_in_0: Callable[[None, float, int], float] = lambda x: 0 if x is None else x

        query = sql_get_walking_report_crew_new(f'01.{month}.{p_date[2]}').getvalue().fetchall()

        result = {}

        row_num = 6
        for row in query:
            ws[f'A{row_num}'] = row[0]
            ws[f'A{row_num}'].style = cell
            ws[f'B{row_num}'] = row[1]
            ws[f'B{row_num}'].style = cell
            ws[f'C{row_num}'] = row[2].strftime('%d.%m.%Y')
            ws[f'C{row_num}'].style = cell
            ws[f'D{row_num}'] = row[3]
            ws[f'D{row_num}'].style = cell
            ws[f'E{row_num}'] = None if row[3] == 1 else 7
            ws[f'E{row_num}'].style = cell
            ws[f'F{row_num}'] = 11 - none_in_0(row[13])
            ws[f'F{row_num}'].style = cell
            ws[f'G{row_num}'] = row[6]
            ws[f'G{row_num}'].style = cell
            ws[f'H{row_num}'] = row[7]
            ws[f'H{row_num}'].style = cell
            ws[f'I{row_num}'] = row[8]
            ws[f'I{row_num}'].style = cell
            ws[f'J{row_num}'] = row[9]
            ws[f'J{row_num}'].style = cell
            ws[f'K{row_num}'] = row[10]
            ws[f'K{row_num}'].style = cell
            ws[f'L{row_num}'] = row[11]
            ws[f'L{row_num}'].style = cell
            ws[f'M{row_num}'] = row[12]
            ws[f'M{row_num}'].style = cell
            ws[f'N{row_num}'] = row[13]
            ws[f'N{row_num}'].style = cell
            row_num += 1
            if result.get(row[1]) is None:
                result[row[1]] = [
                    1,  # количество смен
                    0 if row[3] == 1 else 1,  # количество ночных смен
                    0 if row[3] == 1 else 7,  # количество ночных часов
                    11 - float(none_in_0(row[13])),  # Время переэкскавации
                    none_in_0(row[6]),  # объем переэкскавации
                    none_in_0(row[7]),  # объем повторной переэкскавации
                    none_in_0(row[8]),  # время перегона
                    none_in_0(row[9]),  # время на ТО
                    none_in_0(row[10]),  # аварийный ремонт
                    none_in_0(row[11]),  # время на ППР
                    none_in_0(row[12]),  # эл/часть
                    none_in_0(row[13])  # другие простои
                ]
            else:
                var_res = result[row[1]]
                var_res[0] += 1
                var_res[1] += 0 if row[3] == 1 else 1
                var_res[2] = var_res[1] * 7
                var_res[3] += 11 - none_in_0(row[13])
                var_res[4] += row[6]
                var_res[5] += none_in_0(row[7])
                var_res[6] += none_in_0(row[8])
                var_res[7] += none_in_0(row[9])
                var_res[8] += none_in_0(row[10])
                var_res[9] += none_in_0(row[11])
                var_res[10] += none_in_0(row[12])
                var_res[11] += none_in_0(row[13])
                result[row[1]] = var_res

        row_num = 6

        ws = wb[str('Экипаж')]
        ws['F3'] = f'{[key for key, val in months_keys if val == month][0]} {p_date[2]} г.'

        for k, v in result.items():
            ws[f'A{row_num}'] = k
            ws[f'A{row_num}'].style = cell
            ws[f'B{row_num}'] = v[0]
            ws[f'B{row_num}'].style = cell
            ws[f'C{row_num}'] = v[1]
            ws[f'C{row_num}'].style = cell
            ws[f'D{row_num}'] = v[2]
            ws[f'D{row_num}'].style = cell
            ws[f'E{row_num}'] = v[3]
            ws[f'E{row_num}'].style = cell
            ws[f'F{row_num}'] = v[4]
            ws[f'F{row_num}'].style = cell
            ws[f'G{row_num}'] = v[5]
            ws[f'G{row_num}'].style = cell
            ws[f'H{row_num}'] = v[6]
            ws[f'H{row_num}'].style = cell
            ws[f'I{row_num}'] = v[7]
            ws[f'I{row_num}'].style = cell
            ws[f'J{row_num}'] = v[8]
            ws[f'J{row_num}'].style = cell
            ws[f'K{row_num}'] = v[9]
            ws[f'K{row_num}'].style = cell
            ws[f'L{row_num}'] = v[10]
            ws[f'L{row_num}'].style = cell
            ws[f'M{row_num}'] = v[11]
            ws[f'M{row_num}'].style = cell
            row_num += 1

        wb.save(response)

        return response
