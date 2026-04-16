import openpyxl

from django.views import View
from django.shortcuts import render
from django.http import HttpResponse
from ..helpers import (sql_get_budget_result_24hours,
                       sql_get_budget_result_shift,
                       insert_technics_result_24hours,
                       insert_technics_result_shift,
                       report_folder)


class GetShiftReportPage(View):

    @staticmethod
    def get(request):
        return render(request, 'get_shift_report_page.html')

    @staticmethod
    def post(request):
        month_post = request.POST.get(f'date')
        shift_post = request.POST.get(f'shift')
        if month_post == 'None' or shift_post == 'None':
            return render(
                request,
                'insert_fail_page.html',
                context={
                    'link': 'get_shift_page',
                    'e_label': 'Ошибка выбора даты или смены',
                    'e': 'Необходимо выбрать дату и смену'
                })
        month = str(month_post).split('-')
        shift = int(shift_post)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="PO6_day_{month_post}.xlsx"'
        wb = openpyxl.load_workbook(f'{report_folder}PO6_day.xlsx')
        ws = wb[str(int(month[2]))]
        if shift == 0:
            ws['A9'] = f'за {month[2]}.{month[1]}.{month[0]} г.'
            for row in sql_get_budget_result_24hours(f'{month[2]}.{month[1]}.{month[0]}').getvalue().fetchall():
                if int(row[1]) == 45:
                    insert_technics_result_24hours(ws, 23, row)
                if int(row[1]) == 28:
                    insert_technics_result_24hours(ws, 26, row)
                if int(row[1]) == 31:
                    insert_technics_result_24hours(ws, 27, row)
                if int(row[1]) == 1:
                    insert_technics_result_24hours(ws, 43, row)
                if int(row[1]) == 272:
                    insert_technics_result_24hours(ws, 45, row)
                if int(row[1]) == 9:
                    insert_technics_result_24hours(ws, 47, row)
                if int(row[1]) == 10:
                    insert_technics_result_24hours(ws, 48, row)
                if int(row[1]) == 186:
                    insert_technics_result_24hours(ws, 50, row)
                if int(row[1]) == 4:
                    insert_technics_result_24hours(ws, 53, row)
                if int(row[1]) == 15:
                    insert_technics_result_24hours(ws, 54, row)
                if int(row[1]) == 58:
                    insert_technics_result_24hours(ws, 56, row)
                if int(row[1]) == 43:
                    insert_technics_result_24hours(ws, 57, row)
                if int(row[1]) == 183:
                    insert_technics_result_24hours(ws, 58, row)
                if int(row[1]) == 283:
                    insert_technics_result_24hours(ws, 59, row)
                if int(row[1]) == 355:
                    insert_technics_result_24hours(ws, 60, row)
                if int(row[1]) == 284:
                    insert_technics_result_24hours(ws, 61, row)
                if int(row[1]) == 285:
                    insert_technics_result_24hours(ws, 62, row)
                if int(row[1]) == 286:
                    insert_technics_result_24hours(ws, 63, row)
        else:
            ws['A9'] = f'за {month[2]}.{month[1]}.{month[0]} г. смена №{shift}'
            for row in sql_get_budget_result_shift(f'{month[2]}.{month[1]}.{month[0]}', shift).getvalue().fetchall():
                if int(row[2]) == 45:
                    insert_technics_result_shift(ws, 23, row)
                if int(row[2]) == 28:
                    insert_technics_result_shift(ws, 26, row)
                if int(row[2]) == 31:
                    insert_technics_result_shift(ws, 27, row)
                if int(row[2]) == 1:
                    insert_technics_result_shift(ws, 43, row)
                if int(row[2]) == 272:
                    insert_technics_result_shift(ws, 45, row)
                if int(row[2]) == 9:
                    insert_technics_result_shift(ws, 47, row)
                if int(row[2]) == 10:
                    insert_technics_result_shift(ws, 48, row)
                if int(row[2]) == 186:
                    insert_technics_result_shift(ws, 50, row)
                if int(row[2]) == 4:
                    insert_technics_result_shift(ws, 53, row)
                if int(row[2]) == 15:
                    insert_technics_result_shift(ws, 54, row)
                if int(row[2]) == 58:
                    insert_technics_result_shift(ws, 56, row)
                if int(row[2]) == 43:
                    insert_technics_result_shift(ws, 57, row)
                if int(row[1]) == 183:
                    insert_technics_result_24hours(ws, 58, row)
                if int(row[1]) == 283:
                    insert_technics_result_24hours(ws, 59, row)
                if int(row[1]) == 355:
                    insert_technics_result_24hours(ws, 60, row)
                if int(row[1]) == 284:
                    insert_technics_result_24hours(ws, 61, row)
                if int(row[1]) == 285:
                    insert_technics_result_24hours(ws, 62, row)
                if int(row[1]) == 286:
                    insert_technics_result_24hours(ws, 63, row)
        wb.save(response)
        return response
