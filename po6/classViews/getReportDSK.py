import openpyxl

from openpyxl.styles import NamedStyle, Font, Border, Side, Alignment
from django.views import View
from django.shortcuts import render
from django.http import HttpResponse
from ..helpers import months_keys, p_date, report_folder, sql_get_dsk_report


class GetReportDSK(View):

    @staticmethod
    def get(request):
        range_years = list(range(2020, int(p_date[2])))[::-1]
        return render(request, 'get_report_DSK.html', {
            'months': months_keys,
            'range_years': range_years,
            'year': int(p_date[2])
        })

    @staticmethod
    def post(request):
        month = request.POST.get('month')
        year = request.POST.get('year')

        td = Side(border_style='dotted')
        cell = NamedStyle(name='cell')
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(top=td, left=td, bottom=td, right=td)
        cell.font = Font(name='Times New Roman', size=12)
        cell1 = NamedStyle(name='cell1')
        cell1.alignment = Alignment(horizontal="center", vertical="center")
        cell1.border = Border(top=td, left=td, bottom=td, right=td)
        cell1.font = Font(name='Times New Roman', size=12, bold=True)

        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = f'attachment; filename="Report_DSK_{month}.{year}.xlsx"'
        wb = openpyxl.load_workbook(f'{report_folder}ReportDSK.xlsx')
        ws = wb[str('Отчет')]
        ws['C6'] = f'{[key for key, val in months_keys if val == month][0]} {year} г.'

        dsk_report = sql_get_dsk_report(f'01.{month}.{year}').getvalue().fetchall()
        num = 10
        moto = 0
        for row_sd in dsk_report:
            ws[f'A{num}'] = row_sd[0].strftime('%d.%m.%Y')
            ws[f'A{num}'].style = cell
            ws[f'B{num}'] = row_sd[1]
            ws[f'B{num}'].style = cell
            ws[f'C{num}'] = row_sd[1] + row_sd[3]
            ws[f'C{num}'].style = cell
            ws[f'D{num}'] = row_sd[3]
            ws[f'D{num}'].style = cell
            ws[f'E{num}'].style = cell
            moto += row_sd[3]
            num += 1

        ws.merge_cells(f'A{num}:C{num}')
        ws[f'A{num}'] = 'Итого:'
        ws[f'A{num}'].style = cell1
        ws[f'B{num}'].style = cell1
        ws[f'C{num}'].style = cell1
        ws[f'D{num}'] = moto
        ws[f'D{num}'].style = cell1
        ws[f'E{num}'].style = cell

        num += 2
        ws.merge_cells(f'A{num}:C{num}')
        ws[f'A{num}'] = 'Представитель АО «Золото Северного Урала»'
        ws[f'E{num}'] = 'Фомин И.Н.'
        num += 2
        ws.merge_cells(f'A{num}:C{num}')
        ws[f'A{num}'] = 'Старший диспетчер УГР'
        num += 3
        ws.merge_cells(f'A{num}:C{num}')
        ws[f'A{num}'] = 'Представитель ООО «ЩСУ»'
        num += 2
        ws.merge_cells(f'A{num}:C{num}')
        ws[f'A{num}'] = 'Главный инженер ООО «ЩСУ»'
        ws[f'E{num}'] = 'Абсалямов С.Г.'

        wb.save(response)
        return response
